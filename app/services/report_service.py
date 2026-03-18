from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.config.app_settings import AppSettings
from app.models.records import ProcessedDocument


DETAIL_HEADERS = [
    "Fecha de proceso",
    "Cliente",
    "Nombre original del archivo",
    "Nombre final del archivo",
    "Ruta original",
    "Ruta final",
    "Tipo de archivo",
    "Método de detección de fecha",
    "Fecha detectada del documento",
    "Año asignado",
    "Mes asignado",
    "Estado del procesamiento",
    "Observaciones / incidencias",
    "Número de factura",
    "Emisor / proveedor",
    "NIF emisor",
    "Base imponible",
    "Cuota IVA",
    "Importe total",
    "Moneda",
    "Campos inferidos",
]


class ReportService:
    def __init__(self, settings: AppSettings) -> None:
        self.settings = settings

    def build_report_path(self, client_folder: Path) -> Path:
        stamp = datetime.now().strftime("%Y-%m-%d %H-%M")
        base_name = f"{self.settings.report_base_name} {stamp}.xlsx"
        candidate = client_folder / base_name
        counter = 2
        while candidate.exists():
            candidate = client_folder / f"{self.settings.report_base_name} {stamp} ({counter}).xlsx"
            counter += 1
        return candidate

    def generate_client_report(self, client_folder: Path, documents: list[ProcessedDocument]) -> Path:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"
        detail_sheet = workbook.create_sheet("Detalle")

        self._populate_summary(summary_sheet, documents)
        self._populate_detail(detail_sheet, documents)

        destination = self.build_report_path(client_folder)
        workbook.save(destination)
        return destination

    def _populate_summary(self, sheet, documents: list[ProcessedDocument]) -> None:
        detection_counter = Counter(doc.detection_method for doc in documents)
        total_detected = sum(1 for doc in documents if doc.detected_date is not None)
        rows = [
            ("Fecha y hora de ejecución", datetime.now()),
            ("Total de archivos procesados", len(documents)),
            ("Total con fecha detectada", total_detected),
            ("Total sin fecha", detection_counter.get("SIN_FECHA", 0)),
        ]
        for method, total in sorted(detection_counter.items()):
            rows.append((f"Total método {method}", total))

        sheet.append(["Métrica", "Valor"])
        for metric, value in rows:
            sheet.append([metric, value])

        header_fill = PatternFill("solid", fgColor="DCE6EE")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 24
        for row in sheet.iter_rows(min_row=2, max_col=2):
            if isinstance(row[1].value, datetime):
                row[1].number_format = "DD/MM/YYYY HH:MM"

    def _populate_detail(self, sheet, documents: list[ProcessedDocument]) -> None:
        sheet.append(DETAIL_HEADERS)
        header_fill = PatternFill("solid", fgColor="DCE6EE")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for document in documents:
            sheet.append(
                [
                    document.processed_at,
                    document.client_name,
                    document.original_name,
                    document.final_name,
                    str(document.original_path),
                    str(document.final_path),
                    document.file_type,
                    document.detection_method,
                    document.detected_date,
                    document.assigned_year,
                    document.assigned_month,
                    document.status,
                    document.notes,
                    document.invoice_data.invoice_number,
                    document.invoice_data.issuer,
                    document.invoice_data.issuer_tax_id,
                    document.invoice_data.taxable_base,
                    document.invoice_data.vat_amount,
                    document.invoice_data.total_amount,
                    document.invoice_data.currency,
                    ", ".join(document.invoice_data.inferred_fields),
                ]
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName="DetalleProcesado", ref=sheet.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        date_columns = {"A", "I"}
        amount_columns = {"Q", "R", "S"}
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 14), 40)
            if letter in date_columns:
                for cell in column_cells[1:]:
                    cell.number_format = "DD/MM/YYYY HH:MM"
            if letter in amount_columns:
                for cell in column_cells[1:]:
                    cell.number_format = '#,##0.00'
